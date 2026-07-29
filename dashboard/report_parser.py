from pathlib import Path
from collections import defaultdict
from datetime import datetime

from openpyxl import load_workbook


class ReportParser:
    """
    Parses Master_Comparison_Report.xlsx
    and prepares dashboard data.
    """

    REQUIRED_SHEETS = [
        "MASTER_SUMMARY",
        "ALL_DIFFERENCES",
        "ALL_MISSING_RECORDS",
        "ALL_DUPLICATES",
        "ALL_ZERO_VALUES",
    ]

    def __init__(self, report_path):

        self.report_path = Path(report_path)

        self.workbook = None

        self.summary_sheet = None
        self.difference_sheet = None
        self.missing_sheet = None
        self.duplicate_sheet = None
        self.zero_sheet = None

    # ==========================================================
    # Validation
    # ==========================================================

    def validate_report(self):

        if not self.report_path.exists():
            raise FileNotFoundError(
                f"Report not found : {self.report_path}"
            )

    def validate_sheets(self):

        missing = [
            sheet
            for sheet in self.REQUIRED_SHEETS
            if sheet not in self.workbook.sheetnames
        ]

        if missing:

            raise ValueError(
                f"Missing worksheet(s): {', '.join(missing)}"
            )

    # ==========================================================
    # Workbook
    # ==========================================================

    def load_workbook(self):

        self.validate_report()

        self.workbook = load_workbook(
            self.report_path,
            data_only=True
        )

        self.validate_sheets()

        self.summary_sheet = self.workbook["MASTER_SUMMARY"]
        self.difference_sheet = self.workbook["ALL_DIFFERENCES"]
        self.missing_sheet = self.workbook["ALL_MISSING_RECORDS"]
        self.duplicate_sheet = self.workbook["ALL_DUPLICATES"]
        self.zero_sheet = self.workbook["ALL_ZERO_VALUES"]

    # ==========================================================
    # Store Summary
    # ==========================================================

    def get_store_summary(self):

        rows = list(
            self.summary_sheet.iter_rows(
                min_row=2,
                values_only=True
            )
        )

        headers = [
            cell.value
            for cell in self.summary_sheet[1]
        ]

        stores = []

        for row in rows:

            if all(value is None for value in row):
                continue    

            stores.append(
                dict(zip(headers, row))
            )

        return stores

    # ==========================================================
    # Summary Metrics
    # ==========================================================

    def get_summary(self, stores):

        total_stores = len(stores)

        passed = sum(
            1
            for store in stores
            if str(store["Status"]).upper() == "PASS"
        )

        failed = total_stores - passed

        total_differences = sum(
            int(store["Differences"] or 0)
            for store in stores
        )

        total_missing = sum(
            int(store["Missing Records"] or 0)
            for store in stores
        )

        total_duplicates = sum(
            int(store["Duplicates"] or 0)
            for store in stores
        )

        total_zero_values = sum(
            int(store["Zero Values"] or 0)
            for store in stores
        )

        accuracy = round(
            (passed / total_stores) * 100,
            2
        ) if total_stores else 0

        return {

            "total_stores": total_stores,

            "passed": passed,

            "failed": failed,

            "accuracy": accuracy,

            "differences": total_differences,

            "missing": total_missing,

            "duplicates": total_duplicates,

            "zero_values": total_zero_values

        }

    # ==========================================================
    # Dynamic Sheet Reader
    # ==========================================================

    def read_sheet_records(self, worksheet):

        if worksheet.max_row <= 1:
            return []

        headers = [
            cell.value
            for cell in worksheet[1]
        ]

        if not any(headers):
            return []

        records = []

        for row in worksheet.iter_rows(
            min_row=2,
            values_only=True
        ):

            if all(value is None for value in row):
                continue

            records.append(
                dict(zip(headers, row))
            )

        return records

    # ==========================================================
    # Store Details
    # ==========================================================

    def get_store_details(self, stores):
        details = defaultdict(
            lambda: {
                "differences": [],
                "missing": [],
                "duplicates": [],
                "zero_values": []
            }
        )

        # ------------------------------------------
        # Store Information
        # ------------------------------------------

        for store in stores:

            store_no = str(
                store.get("Store", "")
            ).strip()

            business_date = str(
                store.get("CB Date")
                or store.get("AC Date")
                or store.get("Date")
                or ""
            ).strip()

            store_key = f"{store_no}|{business_date}"

            if not store_no:
                continue

            report_date = (
                store.get("CB Date")
                or store.get("AC Date")
                or store.get("Date")
                or ""
            )

            report_file = self.store_report_paths.get(
                store_key,
                {}
            )

            details[store_key].update({
                "store": store_no,
                "status": store.get("Status"),
                "integration": store.get("Integration"),
                "cb_date": store.get("CB Date") or store.get("Date"),
                "ac_date": store.get("AC Date") or store.get("Date"),
                "cb_file": store.get("CB File"),
                "ac_file": store.get("AC File"),
                "report_file": report_file.get("excel", ""),
                "pdf_file": report_file.get("pdf", "")
            })

            details[store_key]["validation_failure"] = (
                str(store.get("Status", "")).upper()
                not in ["PASS", "FAIL"]
            )

            details[store_key]["failure_reason"] = (
                store.get("Status")
            )

        differences = self.read_sheet_records(
            self.difference_sheet
        )

        missing = self.read_sheet_records(
            self.missing_sheet
        )

        duplicates = self.read_sheet_records(
            self.duplicate_sheet
        )

        zero_values = self.read_sheet_records(
            self.zero_sheet
        )
                # ------------------------------------------
        # Differences
        # ------------------------------------------

        for record in differences:

            store = str(
                record.get("Store", "")
            ).strip()

            business_date = str(
                record.get("Date", "")
            ).strip()

            store_key = f"{store}|{business_date}"

            if not store:
                continue

            details[store_key]["differences"].append(
                record
            )

        # ------------------------------------------
        # Missing Records
        # ------------------------------------------

        for record in missing:

            store = str(
                record.get("Store", "")
            ).strip()

            business_date = str(
                record.get("Date", "")
            ).strip()

            store_key = f"{store}|{business_date}"

            if not store:
                continue

            details[store_key]["missing"].append(
                record
            )

        # ------------------------------------------
        # Duplicate Records
        # ------------------------------------------

        for record in duplicates:

            store = str(
                record.get("Store", "")
            ).strip()

            business_date = str(
                record.get("Date", "")
            ).strip()

            store_key = f"{store}|{business_date}"

            if not store:
                continue

            details[store_key]["duplicates"].append(
                record
            )

        # ------------------------------------------
        # Zero Values
        # ------------------------------------------

        for record in zero_values:

            store = str(
                record.get("Store", "")
            ).strip()

            business_date = str(
                record.get("Date", "")
            ).strip()

            store_key = f"{store}|{business_date}"

            if not store:
                continue

            details[store_key]["zero_values"].append(
                record
            )

        return dict(details)

    # ==========================================================
    # Chart Data
    # ==========================================================

    def get_chart_data(self, summary):

        return {

            "pass_fail": {

                "labels": [
                    "Passed",
                    "Failed"
                ],

                "values": [
                    summary["passed"],
                    summary["failed"]
                ]

            },

            "issue_distribution": {

                "labels": [
                    "Differences",
                    "Missing",
                    "Duplicates",
                    "Zero Values"
                ],

                "values": [

                    summary["differences"],
                    summary["missing"],
                    summary["duplicates"],
                    summary["zero_values"]

                ]

            }

        }

    # ==========================================================
    # Parser
    # ==========================================================

    def parse(self, store_report_paths=None):
        self.store_report_paths = store_report_paths or {}
        self.load_workbook()

        stores = self.get_store_summary()
        for store in stores:

            status = str(
                store.get("Status", "")
            ).upper()

            store["validation_failure"] = (
                status not in ["PASS", "FAIL"]
            )

            store["failure_reason"] = (
                store.get("Status", "")
            )

        summary = self.get_summary(
            stores
        )

        charts = self.get_chart_data(
            summary
        )

        details = self.get_store_details(stores)

        report_title = "CB → AC Migration Validation Report"

        if stores:

            integration = stores[0].get("Integration")

            if integration:

                report_title = f"{integration} Validation Report"

        report_info = {

            "generated_on": datetime.now().strftime("%d-%b-%Y %I:%M %p"),

            "comparison": stores[0].get("Integration", "") if stores else "",

            "report_title": report_title

        }
        return {

            "stores": stores,

            "summary": summary,

            "charts": charts,

            "details": details,

            "report_info": report_info,

            "report_title": report_title

        }